#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
sheet = wb['Sheet1']

row_id = 1
score = []
for row in sheet:
	if row_id != 1:
		sum = sheet.cell(row = row_id, column = 3).value * 0.3
		sum += sheet.cell(row = row_id, column = 4).value * 0.35
		sum += sheet.cell(row = row_id, column = 5).value * 0.34
		sum += sheet.cell(row = row_id, column = 6).value
		sheet.cell(row = row_id, column = 7).value = sum
		score.append(sum)
	row_id += 1

score.sort(reverse=True)

def move_pointer(ptr):
	if ptr >= 0 and score[ptr + 1] == score[ptr]:
       		temp = score[ptr]
        	while ptr >= 0 and score[ptr] == temp:
                	ptr -= 1
	return ptr

a_pointer = int(len(score) * 0.3) - 1
a_pointer = move_pointer(a_pointer)
b_pointer = int(len(score) * 0.7) - 1
b_pointer = move_pointer(b_pointer)

row_id = 1
for row in sheet:
	if row_id != 1:
		total = sheet.cell(row = row_id, column = 7).value
		if a_pointer >= 0 and score[a_pointer] <= total:
			ptr = (a_pointer + 1) // 2 - 1
			ptr = move_pointer(ptr)	
			if ptr >= 0 and score[ptr] <= total:
				sheet.cell(row = row_id, column = 8).value = "A+"
			else:
				sheet.cell(row = row_id, column = 8).value = "A"
		elif b_pointer >= 0 and score[b_pointer] <= total:
			ptr = a_pointer + ((b_pointer + 1) - (a_pointer + 1)) // 2 
			ptr = move_pointer(ptr)
			if ptr > a_pointer and score[ptr] <= total:
				sheet.cell(row = row_id, column = 8).value = "B+"
			else:
				sheet.cell(row = row_id, column = 8).value = "B"
		else:
			ptr = b_pointer + (len(score) - (b_pointer + 1)) // 2 
			ptr = move_pointer(ptr)
			if ptr > b_pointer and score[ptr] <= total:
                                sheet.cell(row = row_id, column = 8).value = "C+"
			else:
				sheet.cell(row = row_id, column = 8).value = "C"
	row_id += 1

wb.save("student.xlsx")
